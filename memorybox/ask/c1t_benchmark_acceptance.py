"""Safe synthetic acceptance for C1T. No Ollama model calls."""
from __future__ import annotations

import hashlib
import json
import sys
import tempfile
from unittest.mock import patch
from pathlib import Path
from typing import Any

from memorybox.ask.i11a.c1t_benchmark import (
    ChunkParameters,
    RunParameters,
    collect_inventory,
    prepare_parameterized_chunks,
    preflight_benchmark,
    register_canonical_generation,
    run_experiment_matrix,
    run_supervised_benchmark,
    write_default_matrix,
)


def _check(
    name: str,
    condition: bool,
    checks: list[str],
    problems: list[str],
    detail: Any = None,
) -> None:
    checks.append(name)
    if not condition:
        problems.append(f"{name}: {detail}")


def _packet(path: Path, conversations: list[list[tuple[str, str, str]]]) -> str:
    citation_rows = []
    blocks = []
    for ci, turns in enumerate(conversations):
        lines = [f"BEGIN CONVERSATION: synthetic-{ci}", f"grouping: thread-{ci}"]
        for cite, when, body in turns:
            lines.extend(
                [
                    f"{when} — Peggy said: [{cite}]",
                    body,
                    "",
                ]
            )
            citation_rows.append(
                {"cite_as": cite, "sent_at": when, "evidence_id": f"e-{cite}"}
            )
        lines.append("END CONVERSATION")
        blocks.append("\n".join(lines))
    paste = (
        "===== SYSTEM INSTRUCTIONS =====\n"
        "Return JSON. Cite every claim.\n\n"
        "===== USER QUESTION AND EVIDENCE =====\n"
        "Tell me what you know.\n\n"
        "===== TRUSTED EMAIL CONVERSATIONS =====\n\n"
        + "\n\n".join(blocks)
        + "\n"
    )
    path.mkdir(parents=True)
    paste_path = path / "MODEL_PASTE.txt"
    paste_path.write_text(paste, encoding="utf-8", newline="\n")
    digest = hashlib.sha256(paste_path.read_bytes()).hexdigest()
    source_map = {
        "frozen_input_sha256": digest,
        "citations": citation_rows,
        "source_config_fingerprint": "synthetic-v1",
    }
    (path / "SOURCE_MAP.json").write_text(
        json.dumps(source_map, indent=2) + "\n", encoding="utf-8", newline="\n"
    )
    (path / "PREPARATION_REPORT.txt").write_text(
        f"synthetic\nfrozen_input_sha256: {digest}\n",
        encoding="utf-8",
        newline="\n",
    )
    return digest


def run_prove_c1t_benchmark() -> dict[str, Any]:
    checks: list[str] = []
    problems: list[str] = []
    with tempfile.TemporaryDirectory() as temporary:
        root = Path(temporary)
        inventory_result = collect_inventory(
            out_dir=root / "inventory",
            ollama_base_url="http://127.0.0.1:9",
            model="gemma4:26b",
        )
        inventory = inventory_result["inventory"]
        _check(
            "inventory_machine_and_human_reports",
            Path(inventory_result["inventory_json"]).is_file()
            and Path(inventory_result["inventory_report"]).is_file(),
            checks,
            problems,
        )
        _check(
            "inventory_unavailable_is_not_zero",
            inventory["ollama"]["available"] is False
            and inventory["ollama"].get("reason"),
            checks,
            problems,
            inventory["ollama"],
        )
        _check(
            "inventory_never_calls_model",
            inventory["models_called"] is False,
            checks,
            problems,
        )
        _check(
            "inventory_recommends_no_case_when_ollama_unavailable",
            inventory["starting_case_recommendation"]["first_case"] is None,
            checks,
            problems,
            inventory["starting_case_recommendation"],
        )

        source = root / "source"
        _packet(
            source,
            [
                [("email_3", "2024-03-01T00:00:00Z", "late " + "x" * 500)],
                [("email_1", "2024-01-01T00:00:00Z", "early " + "y" * 500)],
                [("email_2", "2024-02-01T00:00:00Z", "middle " + "z" * 500)],
            ],
        )
        registered = register_canonical_generation(
            source_dir=source, generations_dir=root / "generations"
        )
        _check(
            "canonical_generation_hash_locked",
            registered.get("ok") is True
            and registered["manifest"]["internally_consistent"] is True
            and registered["manifest"]["citation_count"] == 3,
            checks,
            problems,
            registered,
        )
        parameters = ChunkParameters(
            target_input_tokens=700,
            hard_input_tokens=1000,
            reserved_output_tokens=200,
            safety_margin_tokens=100,
            num_ctx=1400,
        )
        chunks_a = prepare_parameterized_chunks(
            generation_dir=registered["generation_dir"],
            out_root=root / "chunks-a",
            parameters=parameters,
        )
        chunks_b = prepare_parameterized_chunks(
            generation_dir=registered["generation_dir"],
            out_root=root / "chunks-b",
            parameters=parameters,
        )
        rows_a = chunks_a["manifest"]["chunks"]
        rows_b = chunks_b["manifest"]["chunks"]
        _check(
            "chunking_chronological_and_deterministic",
            [r["time_range"]["start"] for r in rows_a]
            == sorted(r["time_range"]["start"] for r in rows_a)
            and [
                (r["email_ids"], r["sha256"], r["estimated_input_tokens"]) for r in rows_a
            ]
            == [
                (r["email_ids"], r["sha256"], r["estimated_input_tokens"]) for r in rows_b
            ],
            checks,
            problems,
            rows_a,
        )
        _check(
            "ordinary_conversations_remain_atomic",
            all(len(r["conversation_ids"]) >= 1 for r in rows_a)
            and chunks_a["manifest"]["evidence_audit"]["ok"],
            checks,
            problems,
        )

        oversized_source = root / "oversized-source"
        turns = [
            (f"email_{i}", f"2024-01-{i:02d}T00:00:00Z", f"turn-{i} " + "q" * 1000)
            for i in range(1, 9)
        ]
        _packet(oversized_source, [turns])
        oversized_gen = register_canonical_generation(
            source_dir=oversized_source,
            generations_dir=root / "oversized-generations",
        )
        oversized = prepare_parameterized_chunks(
            generation_dir=oversized_gen["generation_dir"],
            out_root=root / "oversized-chunks",
            parameters=ChunkParameters(
                target_input_tokens=900,
                hard_input_tokens=1000,
                reserved_output_tokens=200,
                safety_margin_tokens=100,
                num_ctx=1400,
                overlap_messages=3,
            ),
        )
        overflow_rows = oversized["manifest"]["chunks"]
        declared = {
            eid for row in overflow_rows for eid in row.get("overlap_email_ids") or []
        }
        duplicates = {
            eid
            for eid, count in __import__("collections").Counter(
                eid for row in overflow_rows for eid in row["email_ids"]
            ).items()
            if count > 1
        }
        _check(
            "oversized_split_only_complete_emails_with_declared_overlap",
            oversized.get("ok") is True
            and duplicates == declared
            and oversized["manifest"]["evidence_audit"]["ok"]
            and all(
                2 <= len(row["overlap_email_ids"]) <= 3
                for row in overflow_rows
                if row["overlap_email_ids"]
            ),
            checks,
            problems,
            {"duplicates": duplicates, "declared": declared},
        )
        _check(
            "no_individual_email_body_split",
            all(
                sum(
                    1
                    for row in overflow_rows
                    if eid in row.get("primary_email_ids", [])
                )
                == 1
                for eid, _, _ in turns
            ),
            checks,
            problems,
        )
        _check(
            "every_chunk_respects_hard_target_unless_single_email_unavoidable",
            all(
                int(row["estimated_input_tokens"]) <= 1000
                or "single_email_unavoidable_oversize" in row.get("overflow", [])
                for row in overflow_rows
            ),
            checks,
            problems,
            [
                (row["chunk_index"], row["estimated_input_tokens"], row.get("overflow"))
                for row in overflow_rows
            ],
        )
        try:
            ChunkParameters(
                target_input_tokens=1000,
                hard_input_tokens=2000,
                reserved_output_tokens=1000,
                safety_margin_tokens=1000,
                num_ctx=3000,
            ).validate()
            invalid_budget_refused = False
        except ValueError:
            invalid_budget_refused = True
        _check(
            "chunk_context_budget_refused",
            invalid_budget_refused,
            checks,
            problems,
        )

        # Preflight rejects budget before reaching unreachable synthetic Ollama.
        inventory_path = Path(inventory_result["inventory_json"])
        row = rows_a[0]
        preflight_bad = preflight_benchmark(
            chunk_dir=chunks_a["chunk_dir"],
            chunk_index=row["chunk_index"],
            expected_chunk_hash=row["sha256"],
            inventory_path=inventory_path,
            parameters=RunParameters(
                num_ctx=100,
                num_predict=50,
                think=False,
                safety_margin_tokens=50,
                hard_timeout_seconds=1,
                heartbeat_seconds=1,
            ),
            ollama_base_url="http://127.0.0.1:9",
        )
        _check(
            "preflight_budget_fails_before_inference",
            preflight_bad.get("error") == "context_budget_exceeded",
            checks,
            problems,
            preflight_bad,
        )
        valid_preflight_params = RunParameters(
            num_ctx=2048,
            num_predict=128,
            think=False,
            safety_margin_tokens=64,
            hard_timeout_seconds=1,
            heartbeat_seconds=1,
        )
        preflight_cases = []
        for fake_ollama in (
            {"available": False, "reason": "synthetic_unreachable"},
            {
                "available": True,
                "models": [],
                "requested_model": None,
                "running_models": [],
            },
            {
                "available": True,
                "models": [{"name": "gemma4:26b"}],
                "requested_model": {"name": "gemma4:26b"},
                "requested_model_show": {},
                "running_models": [{"name": "gemma4:26b"}],
            },
        ):
            with patch(
                "memorybox.ask.i11a.c1t_benchmark._ollama_inventory",
                return_value=fake_ollama,
            ):
                preflight_cases.append(
                    preflight_benchmark(
                        chunk_dir=chunks_a["chunk_dir"],
                        chunk_index=row["chunk_index"],
                        expected_chunk_hash=row["sha256"],
                        inventory_path=inventory_path,
                        parameters=valid_preflight_params,
                        ollama_base_url="http://127.0.0.1:9",
                    ).get("error")
                )
        _check(
            "preflight_rejects_unreachable_missing_and_stale_model",
            preflight_cases
            == ["ollama_unavailable", "requested_model_missing", "stale_active_model"],
            checks,
            problems,
            preflight_cases,
        )

        # Parent timeout owns the worker and leaves an immutable failed run.
        fake_preflight = {
            "ok": True,
            "chunk_path": str(Path(chunks_a["chunk_dir"]) / row["file"]),
            "chunk": row,
            "chunk_sha256": row["sha256"],
            "estimated_input_tokens": row["estimated_input_tokens"],
            "inventory_sha256": inventory["inventory_sha256"],
            "ollama_base_url": "http://127.0.0.1:9",
            "model_metadata": {"name": "synthetic:no-model"},
            "baseline_resources": {},
        }
        timeout_params = RunParameters(
            model="synthetic:no-model",
            num_ctx=2048,
            num_predict=128,
            think=False,
            safety_margin_tokens=64,
            hard_timeout_seconds=1,
            stall_warning_seconds=1,
            heartbeat_seconds=1,
            timeout_grace_seconds=1,
        )
        timed = run_supervised_benchmark(
            preflight=fake_preflight,
            results_root=root / "results",
            experiment_id="synthetic-timeout",
            repetition=1,
            parameters=timeout_params,
            confirm_model_run=True,
            worker_command_factory=lambda _request, _raw, _partial, _thinking: [
                sys.executable,
                "-c",
                "import time; time.sleep(30)",
            ],
        )
        _check(
            "supervisor_hard_timeout_terminates_worker",
            timed["record"]["execution_status"] == "TIMEOUT"
            and timed["record"]["recovery"]["cleanup"].get("worker_terminated") is True,
            checks,
            problems,
            timed["record"]["recovery"],
        )
        _check(
            "timeout_preserves_artifacts_and_workbook",
            (Path(timed["run_dir"]) / "run_record.json").is_file()
            and (root / "results" / "C1T_RESULTS.csv").is_file()
            and (root / "results" / "C1T_RESULTS.xlsx").is_file(),
            checks,
            problems,
        )
        successful = run_supervised_benchmark(
            preflight=fake_preflight,
            results_root=root / "results",
            experiment_id="synthetic-next-run",
            repetition=1,
            parameters=timeout_params,
            confirm_model_run=True,
            worker_command_factory=lambda _request, raw, partial, thinking: [
                sys.executable,
                "-c",
                (
                    "from pathlib import Path; import json; "
                    f"Path({str(partial)!r}).write_text("
                    "'{\"claims\":[{\"text\":\"synthetic [email_3]\"}]}', encoding='utf-8'); "
                    f"Path({str(thinking)!r}).write_text('', encoding='utf-8'); "
                    f"Path({str(raw)!r}).write_text(json.dumps({{"
                    "'done': True, 'prompt_eval_count': 10, "
                    "'prompt_eval_duration': 1000000000, 'eval_count': 4, "
                    "'eval_duration': 1000000000"
                    "})+'\\n', encoding='utf-8')"
                ),
            ],
        )
        _check(
            "successful_run_can_follow_synthetic_timeout",
            successful.get("ok") is True
            and successful["record"]["execution_status"] == "COMPLETE"
            and successful["record"]["benchmark_outcome"] == "PASS"
            and successful["record"]["quality"]["schema_pass"] is True
            and Path(successful["run_dir"]) != Path(timed["run_dir"]),
            checks,
            problems,
            successful.get("record"),
        )
        from openpyxl import load_workbook

        book = load_workbook(root / "results" / "C1T_RESULTS.xlsx", data_only=False)
        formulas = [
            cell.value
            for row_cells in book["Runs"].iter_rows()
            for cell in row_cells
            if isinstance(cell.value, str) and cell.value.startswith("=HYPERLINK")
        ]
        _check(
            "workbook_uses_relative_hyperlinks",
            len(formulas) == 12
            and any("synthetic-timeout-" in formula for formula in formulas)
            and any("synthetic-next-run-" in formula for formula in formulas),
            checks,
            problems,
            formulas,
        )
        matrix = write_default_matrix(root / "matrix.json")
        _check(
            "matrix_is_sequential_parameterized_a_to_f",
            matrix["execution_concurrency"] == 1
            and [c["case_id"] for c in matrix["cases"]] == list("ABCDEF")
            and all(c["run_parameters"]["temperature"] == 0.1 for c in matrix["cases"]),
            checks,
            problems,
        )
        matrix_refused = run_experiment_matrix(
            matrix_path=root / "matrix.json",
            results_root=root / "never-run",
            chunk_dir=chunks_a["chunk_dir"],
            inventory_path=inventory_path,
            selected_cases={"A"},
            confirm_model_run=False,
            ollama_base_url="http://127.0.0.1:9",
        )
        _check(
            "matrix_requires_explicit_model_confirmation",
            matrix_refused.get("error") == "explicit_confirmation_required"
            and not (root / "never-run").exists(),
            checks,
            problems,
            matrix_refused,
        )

    from memorybox.ask.c1t_gate_acceptance import run_prove_c1t_gate

    gate_payload = run_prove_c1t_gate()
    checks.extend(gate_payload["checks"])
    if gate_payload.get("problems"):
        problems.extend(gate_payload["problems"])

    return {
        "ok": not problems,
        "prove": "c1t_benchmark",
        "checks": checks,
        "problems": problems,
        "model_calls": 0,
    }
